from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Query, BackgroundTasks
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone
from pathlib import Path
import io
import csv
import openpyxl
from openpyxl import Workbook
import time
import random
import math

import models
import auth
from database import get_db, SessionLocal
from config import settings

router = APIRouter(prefix="/api/contatos", tags=["Contatos"])


# ── Schemas ──────────────────────────────────────────────────────────────────

class ContactCreate(BaseModel):
    phone: str
    name: Optional[str] = None
    tags: Optional[str] = None


class ContactUpdate(BaseModel):
    name: Optional[str] = None
    tags: Optional[str] = None
    is_blacklisted: Optional[bool] = None


class ContactOut(BaseModel):
    id: int
    phone: str
    name: Optional[str]
    is_blacklisted: bool
    tags: Optional[str]
    group_score: Optional[int] = None
    created_at: Optional[datetime]

    class Config:
        from_attributes = True


class ContactListOut(BaseModel):
    total: int
    page: int
    page_size: int
    items: List[ContactOut]


class BulkIdsRequest(BaseModel):
    contact_ids: List[int]


class BulkBlacklistRequest(BaseModel):
    contact_ids: List[int]
    blacklist: bool = True


# ── Helpers ──────────────────────────────────────────────────────────────────

def normalize_phone(phone: str) -> str:
    digits = "".join(c for c in phone if c.isdigit())
    if digits.startswith("0"):
        digits = digits[1:]
    if not digits.startswith("55") and len(digits) <= 11:
        digits = "55" + digits
    return digits


def _contact_listas(contact_id: int, db: Session) -> List[dict]:
    rows = (
        db.query(models.Lista)
        .join(models.ContatoLista, models.ContatoLista.lista_id == models.Lista.id)
        .filter(models.ContatoLista.contato_id == contact_id)
        .all()
    )
    return [{"id": r.id, "nome": r.nome, "cor": r.cor} for r in rows]


# ── Score: grupos em comum WhatsApp ──────────────────────────────────────────

# Rastreia jobs de scoring em memória: user_id -> estado
_score_jobs: dict = {}


def _pick_session_for_scoring(user_id: int, db) -> "models.WhatsAppSession | None":
    """Retorna a sessão conectada com menor carga que NÃO está em campanha ativa."""
    try:
        active_sess_ids = (
            db.query(models.CampaignSession.session_id)
            .join(models.Campaign, models.Campaign.id == models.CampaignSession.campaign_id)
            .filter(
                models.Campaign.user_id == user_id,
                models.Campaign.status == models.CampaignStatus.running,
            )
            .subquery()
        )
        return (
            db.query(models.WhatsAppSession)
            .filter(
                models.WhatsAppSession.user_id == user_id,
                models.WhatsAppSession.status == models.SessionStatus.connected,
                ~models.WhatsAppSession.id.in_(active_sess_ids),
            )
            .order_by(models.WhatsAppSession.messages_sent_today.asc())
            .first()
        )
    except Exception:
        return None


def _calcular_score_worker(user_id: int) -> None:
    """Background task: calcula grupo_score para cada contato ainda sem score."""
    import httpx as _httpx

    db = SessionLocal()
    try:
        contacts = (
            db.query(models.Contact)
            .filter(
                models.Contact.user_id == user_id,
                models.Contact.group_score.is_(None),
                models.Contact.is_blacklisted == False,
                models.Contact.is_invalid == False,
            )
            .all()
        )

        _score_jobs[user_id] = {
            "running": True,
            "total": len(contacts),
            "done": 0,
            "errors": 0,
            "error_msg": None,
        }

        if not contacts:
            _score_jobs[user_id]["running"] = False
            return

        session = _pick_session_for_scoring(user_id, db)
        if not session:
            _score_jobs[user_id]["running"] = False
            _score_jobs[user_id]["error_msg"] = "Nenhuma sessão conectada disponível"
            return

        headers = {}
        if settings.WAHA_API_KEY:
            headers["X-Api-Key"] = settings.WAHA_API_KEY

        waha_url = settings.WAHA_API_URL.rstrip("/")
        sess_id = session.session_id

        for batch_start in range(0, len(contacts), 10):
            batch = contacts[batch_start: batch_start + 10]

            for contact in batch:
                try:
                    phone_id = f"{contact.phone}@c.us"
                    url = f"{waha_url}/api/{sess_id}/contacts/{phone_id}/common-groups"
                    resp = _httpx.get(url, headers=headers, timeout=8)

                    if resp.status_code == 200:
                        data = resp.json()
                        if isinstance(data, list):
                            count = len(data)
                        elif isinstance(data, dict):
                            count = data.get("count") or len(data.get("groups", []))
                        else:
                            count = 0
                    elif resp.status_code in (404, 501, 405):
                        # Endpoint não suportado pela versão do WAHA — registra 0
                        count = 0
                    else:
                        count = 0

                    contact.whatsapp_common_groups = count
                    contact.group_score = count
                    contact.score_calculado_em = datetime.now(timezone.utc)
                    print(f"📊 Score calculado: {contact.phone} → {count} grupos em comum")
                    _score_jobs[user_id]["done"] += 1

                except Exception as exc:
                    _score_jobs[user_id]["errors"] += 1
                    print(f"❌ Erro ao calcular score de {contact.phone}: {exc}")

                # Delay gaussiano anti-ban (média 3s, σ 1s, mínimo 1s)
                delay = max(1.0, random.gauss(3.0, 1.0))
                time.sleep(delay)

            # Commit ao final de cada batch
            try:
                db.commit()
            except Exception:
                db.rollback()

    except Exception as exc:
        print(f"❌ Erro fatal no worker de scores: {exc}")
    finally:
        if user_id in _score_jobs:
            _score_jobs[user_id]["running"] = False
        db.close()


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/stats")
def get_stats(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Estatísticas rápidas da base de contatos do usuário."""
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    base = db.query(models.Contact).filter(models.Contact.user_id == current_user.id)
    total = base.count()
    com_nome = base.filter(
        models.Contact.name.isnot(None),
        models.Contact.name != "",
    ).count()
    hoje = base.filter(models.Contact.created_at >= today_start).count()
    blacklist = base.filter(models.Contact.is_blacklisted == True).count()
    return {
        "total": total,
        "com_nome": com_nome,
        "sem_nome": total - com_nome,
        "hoje": hoje,
        "blacklist": blacklist,
    }


@router.post("/calcular-scores")
def iniciar_calculo_scores(
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Inicia cálculo de score via grupos em comum do WAHA em background."""
    job = _score_jobs.get(current_user.id, {})
    if job.get("running"):
        return {
            "status": "already_running",
            "total": job["total"],
            "done": job["done"],
            "estimativa": None,
        }

    total = (
        db.query(models.Contact)
        .filter(
            models.Contact.user_id == current_user.id,
            models.Contact.group_score.is_(None),
            models.Contact.is_blacklisted == False,
            models.Contact.is_invalid == False,
        )
        .count()
    )

    if total == 0:
        return {"status": "no_contacts", "total": 0, "estimativa": "0 segundos"}

    # Estimativa: ~3.5s por contato (gaussiana média 3s + overhead)
    segundos = total * 3.5
    if segundos < 60:
        estimativa = f"{int(segundos)} segundos"
    elif segundos < 3600:
        estimativa = f"{int(segundos / 60)} minutos"
    else:
        estimativa = f"{segundos / 3600:.1f} horas"

    background_tasks.add_task(_calcular_score_worker, current_user.id)
    return {"status": "started", "total": total, "estimativa": estimativa}


@router.get("/calcular-scores/status")
def status_calculo_scores(
    current_user: models.User = Depends(auth.get_current_user),
):
    """Retorna progresso do job de cálculo de scores em andamento."""
    job = _score_jobs.get(current_user.id)
    if not job:
        return {"running": False, "done": 0, "total": 0, "errors": 0}
    return {
        "running": job.get("running", False),
        "done": job.get("done", 0),
        "total": job.get("total", 0),
        "errors": job.get("errors", 0),
        "error_msg": job.get("error_msg"),
    }


@router.post("/importar-csv/preview")
async def preview_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Lê um CSV e retorna headers + primeiras 5 linhas para mapeamento de colunas."""
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.reader(io.StringIO(text))
    headers = []
    preview = []
    for i, row in enumerate(reader):
        if i == 0:
            headers = [h.strip() for h in row]
        else:
            preview.append([str(v).strip() for v in row])
            if len(preview) >= 5:
                break

    if not headers:
        raise HTTPException(status_code=400, detail="CSV vazio ou sem cabeçalho")

    return {"headers": headers, "preview": preview, "filename": file.filename}


@router.post("/importar-csv")
async def import_csv_with_mapping(
    file: UploadFile = File(...),
    col_phone: str = Query(..., description="Nome da coluna de telefone"),
    col_name: Optional[str] = Query(None, description="Nome da coluna de nome"),
    lista_id: Optional[int] = Query(None, description="ID da lista para adicionar os contatos"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Importa contatos de CSV com mapeamento de colunas explícito."""
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    # Validate lista ownership
    lista = None
    if lista_id:
        lista = db.query(models.Lista).filter(
            models.Lista.id == lista_id,
            models.Lista.user_id == current_user.id,
        ).first()
        if not lista:
            raise HTTPException(status_code=404, detail="Lista não encontrada")

    reader = csv.DictReader(io.StringIO(text))
    fieldnames = [f.strip() for f in (reader.fieldnames or [])]

    if col_phone not in fieldnames:
        raise HTTPException(status_code=400, detail=f"Coluna '{col_phone}' não encontrada. Colunas: {fieldnames}")

    imported = skipped = invalid = 0
    errors = []
    new_contact_ids = []

    for row_idx, raw_row in enumerate(reader, start=2):
        row = {k.strip(): v for k, v in raw_row.items()}
        phone_raw = (row.get(col_phone) or "").strip()
        name = (row.get(col_name) or "").strip() or None if col_name else None

        if not phone_raw:
            skipped += 1
            continue

        try:
            phone = normalize_phone(phone_raw)
            if len(phone) < 12:
                errors.append(f"Linha {row_idx}: inválido ({phone_raw})")
                invalid += 1
                continue

            existing = db.query(models.Contact).filter(
                models.Contact.user_id == current_user.id,
                models.Contact.phone == phone,
            ).first()

            if existing:
                if lista and not db.query(models.ContatoLista).filter(
                    models.ContatoLista.contato_id == existing.id,
                    models.ContatoLista.lista_id == lista_id,
                ).first():
                    db.add(models.ContatoLista(contato_id=existing.id, lista_id=lista_id))
                skipped += 1
                continue

            contact = models.Contact(user_id=current_user.id, phone=phone, name=name)
            db.add(contact)
            db.flush()
            new_contact_ids.append(contact.id)
            if lista:
                db.add(models.ContatoLista(contato_id=contact.id, lista_id=lista_id))
            imported += 1

        except Exception as e:
            errors.append(f"Linha {row_idx}: {e}")
            invalid += 1

    db.commit()
    return {
        "imported": imported,
        "skipped": skipped,
        "invalid": invalid,
        "errors": errors[:20],
        "lista_id": lista_id,
    }


@router.delete("/bulk")
def bulk_delete(
    data: BulkIdsRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    deleted = (
        db.query(models.Contact)
        .filter(
            models.Contact.id.in_(data.contact_ids),
            models.Contact.user_id == current_user.id,
        )
        .delete(synchronize_session=False)
    )
    db.commit()
    return {"deleted": deleted}


@router.post("/blacklist/bulk")
def bulk_blacklist(
    data: BulkBlacklistRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    updated = (
        db.query(models.Contact)
        .filter(
            models.Contact.id.in_(data.contact_ids),
            models.Contact.user_id == current_user.id,
        )
        .update({"is_blacklisted": data.blacklist}, synchronize_session=False)
    )
    db.commit()
    return {"updated": updated}


@router.get("", response_model=ContactListOut)
def list_contacts(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    search: Optional[str] = None,
    blacklisted: Optional[bool] = None,
    ddd: Optional[str] = None,
    lista_id: Optional[int] = None,
    sem_lista: Optional[bool] = None,
    data_inicio: Optional[str] = None,
    data_fim: Optional[str] = None,
    min_score: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    query = db.query(models.Contact).filter(models.Contact.user_id == current_user.id)

    if search:
        query = query.filter(
            (models.Contact.phone.ilike(f"%{search}%")) |
            (models.Contact.name.ilike(f"%{search}%"))
        )
    if blacklisted is not None:
        query = query.filter(models.Contact.is_blacklisted == blacklisted)
    if ddd:
        # Phone stored as 55DDXXXXXXXXX — DDD is chars [2:4]
        query = query.filter(models.Contact.phone.like(f"55{ddd}%"))
    if lista_id is not None:
        query = query.join(
            models.ContatoLista,
            models.ContatoLista.contato_id == models.Contact.id,
        ).filter(models.ContatoLista.lista_id == lista_id)
    elif sem_lista:
        sub = (
            db.query(models.ContatoLista.contato_id)
            .filter(
                models.ContatoLista.lista_id.in_(
                    db.query(models.Lista.id).filter(models.Lista.user_id == current_user.id)
                )
            )
            .subquery()
        )
        query = query.filter(~models.Contact.id.in_(sub))
    if data_inicio:
        try:
            query = query.filter(models.Contact.created_at >= datetime.fromisoformat(data_inicio))
        except ValueError:
            pass
    if data_fim:
        try:
            query = query.filter(models.Contact.created_at <= datetime.fromisoformat(data_fim))
        except ValueError:
            pass
    if min_score is not None:
        query = query.filter(
            models.Contact.group_score.isnot(None),
            models.Contact.group_score >= min_score,
        )

    query = query.order_by(
        models.Contact.group_score.desc().nullslast(),
        models.Contact.created_at.desc(),
    )
    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()

    return ContactListOut(total=total, page=page, page_size=page_size, items=items)


@router.post("", response_model=ContactOut, status_code=status.HTTP_201_CREATED)
def create_contact(
    data: ContactCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    phone = normalize_phone(data.phone)
    existing = db.query(models.Contact).filter(
        models.Contact.user_id == current_user.id,
        models.Contact.phone == phone,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="Contato já cadastrado")

    contact = models.Contact(
        user_id=current_user.id,
        phone=phone,
        name=data.name,
        tags=data.tags,
    )
    db.add(contact)
    db.commit()
    db.refresh(contact)
    return contact


@router.get("/exportar/xlsx")
def export_contacts(
    lista_id: Optional[int] = None,
    blacklisted: Optional[bool] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    query = db.query(models.Contact).filter(models.Contact.user_id == current_user.id)
    if search:
        query = query.filter(
            (models.Contact.phone.ilike(f"%{search}%")) |
            (models.Contact.name.ilike(f"%{search}%"))
        )
    if blacklisted is not None:
        query = query.filter(models.Contact.is_blacklisted == blacklisted)
    if lista_id is not None:
        query = query.join(
            models.ContatoLista,
            models.ContatoLista.contato_id == models.Contact.id,
        ).filter(models.ContatoLista.lista_id == lista_id)

    contacts = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Contatos"
    ws.append(["Telefone", "Nome", "Blacklist", "Tags", "Criado em"])

    for c in contacts:
        ws.append([
            c.phone,
            c.name or "",
            "Sim" if c.is_blacklisted else "Não",
            c.tags or "",
            c.created_at.strftime("%d/%m/%Y %H:%M") if c.created_at else "",
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contatos.xlsx"},
    )


@router.get("/backup/info")
def get_backup_info(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Retorna info do backup mais recente disponível para download."""
    backup = db.query(models.ContactBackup).filter(
        models.ContactBackup.user_id == current_user.id
    ).order_by(models.ContactBackup.created_at.desc()).first()

    if not backup:
        return {"available": False}

    return {
        "available": True,
        "filename": backup.filename,
        "contact_count": backup.contact_count,
        "created_at": backup.created_at,
    }


@router.get("/backup/download")
def download_backup(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Baixa o backup CSV mais recente de contatos."""
    backup = db.query(models.ContactBackup).filter(
        models.ContactBackup.user_id == current_user.id
    ).order_by(models.ContactBackup.created_at.desc()).first()

    if not backup:
        raise HTTPException(status_code=404, detail="Nenhum backup disponível")

    filepath = Path("backups") / str(current_user.id) / backup.filename
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="Arquivo de backup não encontrado")

    return FileResponse(
        path=str(filepath),
        filename=backup.filename,
        media_type="text/csv",
    )


@router.get("/{contact_id}")
def get_contact(
    contact_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    contact = db.query(models.Contact).filter(
        models.Contact.id == contact_id,
        models.Contact.user_id == current_user.id,
    ).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contato não encontrado")

    listas = _contact_listas(contact_id, db)

    # Campanhas recentes enviadas para este contato
    campanhas = (
        db.query(models.CampaignContact)
        .filter(models.CampaignContact.contact_id == contact_id)
        .order_by(models.CampaignContact.sent_at.desc())
        .limit(5)
        .all()
    )
    camp_out = []
    for cc in campanhas:
        camp = cc.campaign
        if camp:
            camp_out.append({
                "id": camp.id,
                "nome": camp.name,
                "status": cc.status,
                "sent_at": cc.sent_at,
            })

    return {
        "id": contact.id,
        "phone": contact.phone,
        "name": contact.name,
        "is_blacklisted": contact.is_blacklisted,
        "tags": contact.tags,
        "created_at": contact.created_at,
        "listas": listas,
        "campanhas_recentes": camp_out,
    }


@router.put("/{contact_id}", response_model=ContactOut)
def update_contact(
    contact_id: int,
    data: ContactUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    contact = db.query(models.Contact).filter(
        models.Contact.id == contact_id,
        models.Contact.user_id == current_user.id,
    ).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contato não encontrado")

    if data.name is not None:
        contact.name = data.name
    if data.tags is not None:
        contact.tags = data.tags
    if data.is_blacklisted is not None:
        contact.is_blacklisted = data.is_blacklisted

    db.commit()
    db.refresh(contact)
    return contact


@router.delete("/{contact_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_contact(
    contact_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    contact = db.query(models.Contact).filter(
        models.Contact.id == contact_id,
        models.Contact.user_id == current_user.id,
    ).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contato não encontrado")
    db.delete(contact)
    db.commit()


@router.post("/{contact_id}/blacklist")
def toggle_blacklist(
    contact_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    contact = db.query(models.Contact).filter(
        models.Contact.id == contact_id,
        models.Contact.user_id == current_user.id,
    ).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Contato não encontrado")

    contact.is_blacklisted = not contact.is_blacklisted
    db.commit()
    return {"phone": contact.phone, "is_blacklisted": contact.is_blacklisted}


@router.post("/importar")
async def import_contacts(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Legacy: importa contatos via XLSX (col A = telefone, col B = nome)."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Arquivo deve ser .xlsx ou .xls")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    imported = skipped = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        phone_raw = str(row[0]).strip()
        name = str(row[1]).strip() if len(row) > 1 and row[1] else None

        try:
            phone = normalize_phone(phone_raw)
            if len(phone) < 10:
                errors.append(f"Linha {row_idx}: número inválido ({phone_raw})")
                skipped += 1
                continue

            existing = db.query(models.Contact).filter(
                models.Contact.user_id == current_user.id,
                models.Contact.phone == phone,
            ).first()

            if existing:
                skipped += 1
                continue

            db.add(models.Contact(user_id=current_user.id, phone=phone, name=name))
            imported += 1
        except Exception as e:
            errors.append(f"Linha {row_idx}: {str(e)}")
            skipped += 1

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors[:20]}


@router.post("/upload")
async def upload_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user),
):
    """Legacy: importa contatos via CSV."""
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Arquivo deve ser .csv")

    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    fieldnames = [f.strip().lower() for f in (reader.fieldnames or [])]

    col_phone = next((f for f in fieldnames if f in ("telefone", "phone", "numero", "número")), None)
    col_name = next((f for f in fieldnames if f in ("nome", "name")), None)

    if col_phone is None:
        raise HTTPException(
            status_code=400,
            detail=f"CSV deve ter coluna 'telefone' ou 'phone'. Colunas encontradas: {fieldnames}"
        )

    imported = skipped = 0
    errors = []

    for row_idx, raw_row in enumerate(reader, start=2):
        row = {k.strip().lower(): v for k, v in raw_row.items()}
        phone_raw = (row.get(col_phone) or "").strip()
        name = (row.get(col_name) or "").strip() or None

        if not phone_raw:
            skipped += 1
            continue

        try:
            phone = normalize_phone(phone_raw)
            if len(phone) < 10:
                errors.append(f"Linha {row_idx}: número inválido ({phone_raw})")
                skipped += 1
                continue

            existing = db.query(models.Contact).filter(
                models.Contact.user_id == current_user.id,
                models.Contact.phone == phone,
            ).first()

            if existing:
                skipped += 1
                continue

            db.add(models.Contact(user_id=current_user.id, phone=phone, name=name))
            imported += 1
        except Exception as e:
            errors.append(f"Linha {row_idx}: {e}")
            skipped += 1

    db.commit()
    return {"imported": imported, "skipped": skipped, "errors": errors[:20]}
